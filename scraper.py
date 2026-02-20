"""
Facebook Group Post Scraper
============================
Library : Playwright (async) — konek ke Chrome yang sudah login via Remote Debug
Output  : JSON + Excel (.xlsx)
Akun FB : bree420925@gmail.com

Cara kerja:
  - Jika Chrome sudah terbuka → script konek langsung (tidak ganggu sesi aktif)
  - Jika Chrome belum terbuka → script buka Chrome otomatis dengan profil yang benar
  - Tidak perlu tutup Chrome dulu, tidak perlu input email/password

Install:
    pip install playwright openpyxl requests
    # TIDAK perlu: playwright install chromium

Usage:
    python scraper.py
"""

import asyncio
import json
import re
import sys
import random
import subprocess
import socket
import time
import requests
from datetime import datetime
from pathlib import Path

from playwright.async_api import async_playwright, TimeoutError as PWTimeout
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import platform, os

def _get_chrome_profile_path() -> str:
    """Deteksi otomatis path Chrome User Data sesuai OS."""
    system = platform.system()
    home = Path.home()
    if system == "Windows":
        base = Path(os.environ.get("LOCALAPPDATA", home / "AppData/Local"))
        return str(base / "Google/Chrome/User Data")
    elif system == "Darwin":  # macOS
        return str(home / "Library/Application Support/Google/Chrome")
    else:  # Linux
        return str(home / ".config/google-chrome")

def _get_chrome_executable() -> str:
    """Deteksi path executable Chrome sesuai OS."""
    system = platform.system()
    candidates = {
        "Windows": [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        ],
        "Darwin": [
            "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
        ],
        "Linux": [
            "/usr/bin/google-chrome",
            "/usr/bin/chromium-browser",
            "/usr/bin/chromium",
        ],
    }
    for path in candidates.get(system, []):
        if Path(path).exists():
            return path
    return "google-chrome"  


CONFIG = {
   
    "chrome_user_data_dir": r"C:\Users\KuroX66\AppData\Local\Google\Chrome\User Data",
    "chrome_profile":       "Default",   
    "chrome_executable":    r"C:\Program Files\Google\Chrome\Application\chrome.exe",
    "debug_port":           9222,

    "groups": [
        {
            "name": "Grup 1 - HP Second",
            "url": "https://www.facebook.com/groups/513315927639144?locale=id_ID",
            "max_posts": 300
        },
        
         {
            "name": "Grup 2 - Gadget Bekas",
            "url": "https://www.facebook.com/groups/2031638100443573?locale=id_ID",
           "max_posts": 300
     }
    ],

    "scroll_pause": (2.5, 3.5),
    "headless":     False,
    "output_dir":   "output",
    "save_json":    True,
    "save_excel":   True,
}

class PostParser:
    @staticmethod
    def parse(text: str) -> dict:
        result = {
            "barang":           PostParser._extract_barang(text),
            "harga":            PostParser._extract_harga(text),
            "kondisi_baterai":  PostParser._extract_bh(text),
            "garansi":          PostParser._extract_garansi(text),
            "nomor_wa":         PostParser._extract_wa(text),
            "kartu":            PostParser._extract_kartu(text),
            "kelengkapan":      PostParser._extract_kelengkapan(text),
            "tt_bt":            PostParser._extract_ttbt(text),
            "teks_asli":        text.strip(),
        }
        return result

    @staticmethod
    def _extract_barang(text):
        patterns = [
            r"(?:di\s*)?jual\s+(.+?)(?:\n|$)",
            r"(?:^|\n)((?:iPhone|Samsung|Xiaomi|Oppo|Vivo|Realme|Poco|Redmi|HP)\s[^\n]{5,60})",
        ]
        for p in patterns:
            m = re.search(p, text, re.IGNORECASE)
            if m:
                return m.group(1).strip()[:120]
        return ""

    @staticmethod
    def _extract_harga(text):
        patterns = [
            r"harga\s*[:\-]?\s*(rp\.?\s*[\d.,]+)",
            r"(rp\.?\s*[\d.,]{4,})",
            r"harga\s*[:\-]?\s*([\d.,]{5,})",
        ]
        for p in patterns:
            m = re.search(p, text, re.IGNORECASE)
            if m:
                raw = re.sub(r"[Rr][Pp]\.?\s*", "", m.group(1))
                raw = raw.replace(".", "").replace(",", "")
                try:
                    return int(raw)
                except ValueError:
                    pass
        return None

    @staticmethod
    def _extract_bh(text):
        m = re.search(r"(?:bh|battery health)\s*[:\-]?\s*(\d+)\s*%", text, re.IGNORECASE) \
            or re.search(r"(\d+)\s*%\s*(?:bh|masih ori|battery)", text, re.IGNORECASE)
        return int(m.group(1)) if m else None

    @staticmethod
    def _extract_garansi(text):
        if re.search(r"resmi\s*ibox", text, re.IGNORECASE):
            return "Resmi iBox"
        if re.search(r"garansi\s*resmi", text, re.IGNORECASE):
            return "Resmi"
        if re.search(r"garansi\s*distributor", text, re.IGNORECASE):
            return "Distributor"
        if re.search(r"inter\s*garansi|inter national", text, re.IGNORECASE):
            return "Inter"
        return ""

    @staticmethod
    def _extract_wa(text):
        # Pattern 1: Dengan prefix (wa, whatsapp, hub, minat, kontak) — support kurung & dash
        m = re.search(
            r"(?:wa|whatsapp|hub|minat|kontak)\s*[:\-]?\s*((?:08|62|\+62)[\d\s\-()]{8,14})",
            text, re.IGNORECASE
        )
        # Pattern 2: Langsung nomor — jika pattern 1 tidak ketemu
        if not m:
            m = re.search(r"((?:08|62|\+62)[\d]{9,13})", text)
        # Pattern 3: Nomor dengan kurung/dash (tanpa prefix) — sebagai fallback
        if not m:
            m = re.search(r"((?:08|62|\+62)[\d\s\-()]{9,14})", text)
        
        if m:
            # Clean: hapus semua spaces, dashes, dan kurung — tinggal angka & +
            return re.sub(r"[\s\-()\[\]]", "", m.group(1))
        return ""

    @staticmethod
    def _extract_kartu(text):
        if re.search(r"semua\s*kartu", text, re.IGNORECASE):
            return "Semua Kartu"
        if re.search(r"jaringan\s*permanen", text, re.IGNORECASE):
            return "Jaringan Permanen"
        if re.search(r"\bgsm\b", text, re.IGNORECASE):
            return "GSM"
        return ""

    @staticmethod
    def _extract_kelengkapan(text):
        if re.search(r"fullset|full\s*set", text, re.IGNORECASE):
            return "Fullset"
        if re.search(r"dus|box", text, re.IGNORECASE):
            return "Ada Dus"
        return ""

    @staticmethod
    def _extract_ttbt(text):
        if re.search(r"tidak\s*terima\s*tt", text, re.IGNORECASE):
            return "Tidak TT/BT"
        if re.search(r"tt\s*ok|terima\s*tt", text, re.IGNORECASE):
            return "TT OK"
        return ""

class ExcelExporter:
    HEADERS = [
        "Barang", "Harga (Rp)", "BH (%)", "Garansi",
        "No. WA", "Kartu", "Kelengkapan", "TT/BT", "Teks Asli"
    ]
    KEYS = [
        "barang", "harga", "kondisi_baterai", "garansi",
        "nomor_wa", "kartu", "kelengkapan", "tt_bt", "teks_asli"
    ]

    @classmethod
    def export(cls, posts: list[dict], filepath: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hasil Scraping"

        header_fill = PatternFill("solid", fgColor="1E3A5F")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for col, h in enumerate(cls.HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        alt_fill = PatternFill("solid", fgColor="F0F4F8")
        for row_idx, post in enumerate(posts, 2):
            fill = alt_fill if row_idx % 2 == 0 else PatternFill()
            for col_idx, key in enumerate(cls.KEYS, 1):
                val = post.get(key, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=(key == "teks_asli"), vertical="top")

        widths = [35, 15, 10, 15, 18, 18, 15, 12, 60]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        wb.save(filepath)
        print(f"[Excel] Tersimpan: {filepath}")

class FBGroupScraper:
    def __init__(self, config: dict):
        self.config = config
        self.posts: list[dict] = []
        Path(config["output_dir"]).mkdir(exist_ok=True)

    async def run(self):
        async with async_playwright() as p:
            port = self.config["debug_port"]

            if self._is_debug_port_open(port):
                print(f"[Browser] ✅ Chrome terdeteksi di port {port}. Konek langsung...")
                browser = await p.chromium.connect_over_cdp(f"http://localhost:{port}")
            else:
                print(f"[Browser] Chrome belum terbuka dengan debug port.")
                print(f"[Browser] Membuka Chrome dengan debug port {port}...")
                self._launch_chrome_with_debug_port(port)
                await self._wait_for_debug_port(port)
                browser = await p.chromium.connect_over_cdp(f"http://localhost:{port}")

            context = browser.contexts[0] if browser.contexts else await browser.new_context()
            
            print(f"\n{'='*60}")
            print(f"✅ MODE PARALLEL: Buka {len(self.config['groups'])} tab sekaligus")
            print(f"{'='*60}\n")
            
            # ✅ Buat pages untuk semua grup
            pages = []
            for group in self.config["groups"]:
                page = await context.new_page()
                pages.append(page)
            
            # ✅ Verify login 1x (gunakan page pertama)
            await self._verify_login(pages[0])
            
            # ✅ PARALLEL: Scrape semua grup bersamaan
            tasks = []
            for idx, (page, group) in enumerate(zip(pages, self.config["groups"]), 1):
                task = self._scrape_single_group(page, group, idx)
                tasks.append(task)
            
            # Tunggu semua selesai
            results = await asyncio.gather(*tasks)
            
            # Collect semua results
            all_posts = []
            for posts_list in results:
                all_posts.extend(posts_list)
            
            print(f"\n{'='*60}")
            print(f"✅ SELESAI SEMUA GRUP SECARA PARALLEL. Total: {len(all_posts)} posts")
            print(f"{'='*60}\n")
            
            self.posts = all_posts
        
        return self.posts

    async def _scrape_single_group(self, page, group: dict, idx: int) -> list[dict]:
        """Scrape satu grup dalam task terpisah (parallel)."""
        print(f"\n[Tab {idx}] Mulai scraping: {group['name']}")
        
        # Buat instance scraper lokal untuk grup ini (avoid race condition)
        scraper_instance = FBGroupScraper(self.config)
        scraper_instance.posts = []
        scraper_instance.config["group_url"] = group["url"]
        scraper_instance.config["max_posts"] = group["max_posts"]
        
        # Scrape dengan instance lokal
        await scraper_instance._scrape_group(page)
        
        # Save hasil
        saved_posts = scraper_instance.posts.copy()
        scraper_instance._save_results(group["name"])
        
        print(f"[Tab {idx}] ✅ Selesai: {group['name']} ({len(scraper_instance.posts)} posts)")
        
        return saved_posts

    def _is_debug_port_open(self, port: int) -> bool:
        """Cek apakah Chrome sudah berjalan dengan remote debug port."""
        try:
            resp = requests.get(f"http://localhost:{port}/json", timeout=2)
            return resp.status_code == 200
        except Exception:
            return False

    def _launch_chrome_with_debug_port(self, port: int):
        """
        Buka Chrome baru dengan remote debugging port aktif.
        Chrome akan terbuka dengan profil yang sudah login.
        """
        exe = self.config["chrome_executable"]
        user_data = self.config["chrome_user_data_dir"]
        profile = self.config["chrome_profile"]

        if not Path(exe).exists():
            raise FileNotFoundError(
                f"Chrome tidak ditemukan di: {exe}\n"
                f"Edit CONFIG['chrome_executable'] dengan path yang benar."
            )

        cmd = [
            exe,
            f"--remote-debugging-port={port}",
            f"--user-data-dir={user_data}",   
            f"--profile-directory={profile}",
            "--no-first-run",
            "--no-default-browser-check",
            "--disable-blink-features=AutomationControlled",
            "about:blank",  # Buka halaman kosong, URL grup akan di-buka di loop
        ]

        try:
            proc = subprocess.Popen(
                cmd,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                shell=False,
            )
            print(f"[Browser] Chrome diluncurkan (PID: {proc.pid}). Menunggu siap...")
        except Exception as e:
            raise RuntimeError(f"Gagal membuka Chrome: {e}")

    async def _wait_for_debug_port(self, port: int, timeout: int = 30):
        """Tunggu hingga Chrome siap menerima koneksi debug."""
        print(f"[Browser] Menunggu Chrome siap di port {port}...")
        deadline = time.time() + timeout
        attempt = 0
        while time.time() < deadline:
            if self._is_debug_port_open(port):
                print(f"[Browser] ✅ Chrome siap di port {port}.")
                return
            attempt += 1
            if attempt % 6 == 0:  # log tiap 3 detik
                elapsed = int(time.time() - (deadline - timeout))
                print(f"[Browser] Masih menunggu... ({elapsed}s)")
            await asyncio.sleep(0.5)

        print("\n" + "=" * 60)
        print("❌ Chrome tidak merespons. Coba langkah berikut:")
        print()
        print("1. Buka CMD baru, jalankan perintah ini secara manual:")
        print(f'   "{self.config["chrome_executable"]}" --remote-debugging-port={port} --user-data-dir="{self.config["chrome_user_data_dir"]}"')
        print()
        print("2. Lalu buka browser lain dan akses:")
        print(f"   http://localhost:{port}/json")
        print()
        print("3. Jika muncul JSON → jalankan script lagi (Chrome sudah siap)")
        print("   Jika error → Chrome belum support debug port, coba update Chrome")
        print("=" * 60)
        raise TimeoutError(f"Chrome tidak merespons di port {port} setelah {timeout} detik.")

    async def _verify_login(self, page):
        """
        Cek apakah profil Chrome sudah login ke Facebook.
        Jika belum, tampilkan instruksi dan berhenti.
        """
        print("[Login] Memeriksa status login Facebook...")
        await page.goto("https://www.facebook.com/", wait_until="domcontentloaded")
        await self._human_delay(2, 3)

        url = page.url
        if "login" in url or "checkpoint" in url:
            print("=" * 60)
            print("⚠️  Chrome profil ini BELUM login ke Facebook!")
            print()
            print("Cara fix:")
            print("  1. Buka Chrome secara normal")
            print("  2. Login ke Facebook dengan akun bree420925@gmail.com")
            print("  3. Tutup Chrome")
            print("  4. Jalankan script ini lagi")
            print("=" * 60)
            raise Exception("Facebook belum login di profil Chrome ini.")

        try:
            name_el = await page.query_selector("[aria-label='Your profile']")
            name = await name_el.get_attribute("aria-label") if name_el else "—"
        except Exception:
            name = "—"

        print(f"[Login] ✅ Sudah login! Akun terdeteksi: {name}")

    JS_AUTO_EXPAND = """
    (() => {
        // Hindari inject ulang jika sudah aktif
        if (window.__fbAutoExpandActive) return;
        window.__fbAutoExpandActive = true;

        const BUTTON_TEXTS = [
            'See more', 'Lihat selengkapnya', 'See translation',
            'Lihat terjemahan', 'Selengkapnya'
        ];

        const clicked = new WeakSet();

        const isVisible = (el) => {
            if (!(el instanceof HTMLElement)) return false;
            const style = window.getComputedStyle(el);
            if (style.display === 'none' || style.visibility === 'hidden' || style.opacity === '0')
                return false;
            const rect = el.getBoundingClientRect();
            return rect.width > 0 && rect.height > 0;
        };

        const normalizedText = (el) =>
            (el?.innerText || el?.textContent || '').replace(/\\s+/g, ' ').trim();

        const selectors = [
            'div[role="button"]', 'span[role="button"]',
            'a[role="button"]', 'button'
        ];

        const getCandidates = () =>
            selectors.flatMap(s => Array.from(document.querySelectorAll(s)));

        window.__fbClickExpand = () => {
            let count = 0;
            for (const el of getCandidates()) {
                if (!isVisible(el) || clicked.has(el)) continue;
                const text = normalizedText(el);
                if (!text) continue;
                if (BUTTON_TEXTS.some(t => text.toLowerCase() === t.toLowerCase())) {
                    clicked.add(el);
                    el.click();
                    count++;
                }
            }
            return count;
        };

        // Helper ambil teks semua postingan via article
        window.__fbGetVisiblePostTexts = () => {
            const articles = Array.from(document.querySelectorAll('[role="article"]'));
            return articles
                .map(n => normalizedText(n))
                .filter(t => t.length > 20);
        };

        window.__fbAutoExpandStop = () => {
            window.__fbAutoExpandActive = false;
            delete window.__fbAutoExpandStop;
        };
    })();
    """

    async def _scrape_group(self, page):
        print(f"[Scraper] Membuka grup: {self.config['group_url']}")
        await page.goto(self.config["group_url"], wait_until="domcontentloaded")
        await self._human_delay(3, 5)

        try:
            await page.click("[aria-label='Close']", timeout=3000)
        except PWTimeout:
            pass

        await page.evaluate(self.JS_AUTO_EXPAND)
        print("[Scraper] Auto-expand 'See more' aktif.")

        seen_texts = set()
        scroll_attempts = 0
        max_scroll_without_new = 8

        print(f"[Scraper] Target: {self.config['max_posts']} postingan. Mulai scroll...")

        while len(self.posts) < self.config["max_posts"]:

            clicked = await page.evaluate("window.__fbClickExpand ? window.__fbClickExpand() : 0")
            if clicked > 0:
                print(f"[Expand] Klik {clicked} tombol 'See more'")
                await self._human_delay(1.0, 1.8)  
   
            new_found = await self._extract_posts(page, seen_texts)

            if new_found == 0:
                scroll_attempts += 1
                if scroll_attempts >= max_scroll_without_new:
                    print("[Scraper] Tidak ada postingan baru ditemukan. Berhenti.")
                    break
            else:
                scroll_attempts = 0
                print(f"[Scraper] Total terkumpul: {len(self.posts)}/{self.config['max_posts']}")

            await page.evaluate("window.scrollBy(0, window.innerHeight * 2)")
            pause = random.uniform(*self.config["scroll_pause"])
            await asyncio.sleep(pause)

            # 4. Re-inject jika halaman reload/navigate (jaga-jaga)
            await page.evaluate(self.JS_AUTO_EXPAND)

        print(f"[Scraper] Selesai. Total: {len(self.posts)} postingan.")

    async def _extract_posts(self, page, seen_texts: set) -> int:
        """
        Ekstrak teks postingan dari DOM.
        Prioritas: gunakan __fbGetVisiblePostTexts() dari JS inject,
        fallback ke selector CSS biasa.
        """
        new_found = 0

        # Coba ambil via helper JS (lebih akurat karena sudah di-expand)
        try:
            texts: list[str] = await page.evaluate(
                "window.__fbGetVisiblePostTexts ? window.__fbGetVisiblePostTexts() : []"
            )
            for text in texts:
                text = text.strip()
                if len(text) < 20 or text in seen_texts:
                    continue
                if len(self.posts) >= self.config["max_posts"]:
                    break
                seen_texts.add(text)
                parsed = PostParser.parse(text)
                parsed["scraped_at"] = datetime.now().isoformat()
                self.posts.append(parsed)
                new_found += 1

            if new_found > 0:
                return new_found
        except Exception:
            pass

        # Fallback: selector CSS biasa
        selectors = [
            "[data-ad-preview='message']",
            "div[data-testid='post_message']",
            "div.x1iorvi4.x1pi30zi.x1l90r2v.x1swvt13",
        ]
        for selector in selectors:
            try:
                elements = await page.query_selector_all(selector)
                for el in elements:
                    text = (await el.inner_text()).strip()
                    if len(text) < 20 or text in seen_texts:
                        continue
                    if len(self.posts) >= self.config["max_posts"]:
                        break
                    seen_texts.add(text)
                    parsed = PostParser.parse(text)
                    parsed["scraped_at"] = datetime.now().isoformat()
                    self.posts.append(parsed)
                    new_found += 1
            except Exception:
                continue

        return new_found

    def _save_results(self, group_name: str = "posts"):
        """Save hasil scraping dengan nama grup."""
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out = self.config["output_dir"]
        
        # Sanitize group_name untuk filename (replace spasi & karakter aneh)
        safe_name = re.sub(r"[^\w\s-]", "", group_name).strip().replace(" ", "_")
        filename_prefix = f"{safe_name}_{ts}" if group_name != "posts" else f"posts_{ts}"

        if self.config["save_json"]:
            json_path = f"{out}/{filename_prefix}.json"
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(self.posts, f, ensure_ascii=False, indent=2)
            print(f"[JSON] Tersimpan: {json_path}")

        if self.config["save_excel"]:
            xlsx_path = f"{out}/{filename_prefix}.xlsx"
            ExcelExporter.export(self.posts, xlsx_path)
            print(f"[Excel] Tersimpan: {xlsx_path}")

    @staticmethod
    async def _human_delay(min_s: float, max_s: float):
        """Jeda acak seperti manusia."""
        await asyncio.sleep(random.uniform(min_s, max_s))

if __name__ == "__main__":
    scraper = FBGroupScraper(CONFIG)
    asyncio.run(scraper.run())